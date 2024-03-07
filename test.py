"""Программа делает скриншот, преобразует в строковый формат и сохраняет в excel файле"""

from PIL import Image
import pyautogui
import openpyxl
from time import sleep

sleep(3)   # Задержка, чтобы сохранить тот экран, который нужен

# Делаем скриншот всего экрана
screenshot = pyautogui.screenshot()

# Уменьшаем разрешение изображения до 192x108 пикселей
resized_image = screenshot.resize((192, 108), Image.LANCZOS)

resized_image.save('resized_image.png')


# Создаем новый Excel-файл
wb = openpyxl.Workbook()
ws_white = wb.create_sheet('white')
ws_yellow = wb.create_sheet('yellow')
ws_red = wb.create_sheet('red')
ws_violet = wb.create_sheet('violet')
ws_light_blue = wb.create_sheet('light_blue')
ws_green = wb.create_sheet('green')
ws_blue = wb.create_sheet('blue')
ws_gray = wb.create_sheet('gray')
ws_black = wb.create_sheet('black')
ws_all = wb.create_sheet('all')



# Получаем RGB цифры каждого пикселя и записываем в соответствующий массив и Excel-файл
for y in range(resized_image.height):
    for x in range(resized_image.width):
        r, g, b = resized_image.getpixel((x, y))
        rgb_str = f"{r},{g},{b}"
        ws_all.cell(row=y + 1, column=x + 1, value=rgb_str)

        if r >= 188 and g >= 188 and b >= 188:
            ws_white.cell(row=y + 1, column=x + 1, value=1)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r >= 188 and g >= 188 and b < 188:
            # yellow_pixels.append(1)
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=1)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r >= 188 and g < 188 and b >= 188:
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=1)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r >= 188 and g < 188 and b < 188:
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=1)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r < 188 and g >= 188 and b >= 188:
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=1)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r < 188 and g >= 122 and b < 188:
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=1)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r < 188 and g < 188 and b >= 188:
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=1)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r < 188 and g < 188 and (b >= 56 and b < 188):
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=1)
            ws_black.cell(row=y + 1, column=x + 1, value=0)
        elif r < 188 and g < 188 and b < 56:
            ws_white.cell(row=y + 1, column=x + 1, value=0)
            ws_yellow.cell(row=y + 1, column=x + 1, value=0)
            ws_red.cell(row=y + 1, column=x + 1, value=0)
            ws_violet.cell(row=y + 1, column=x + 1, value=0)
            ws_light_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_green.cell(row=y + 1, column=x + 1, value=0)
            ws_blue.cell(row=y + 1, column=x + 1, value=0)
            ws_gray.cell(row=y + 1, column=x + 1, value=0)
            ws_black.cell(row=y + 1, column=x + 1, value=1)

# Сохраняем Excel-файл
wb.save('pixels_colors.xlsx')

# Открываем созданный Excel-файл
wb = openpyxl.load_workbook('pixels_colors.xlsx')

# Изменяем ширину столбцов в каждом листе до значения 3 (26 пикселей)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 3

# Сохраняем изменения
wb.save('pixels_colors.xlsx')

